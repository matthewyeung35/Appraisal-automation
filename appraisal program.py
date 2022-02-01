from pynput.keyboard import Key, Controller
from datetime import datetime
from openpyxl import load_workbook
import time

 # takes a str money value, return with formatted int
def money_format(money):
    if ',' in money:
        return money
    else:
        money = int(money)
        return '{:,}'.format(money)

# date formater, input 2021-01-01, return January 1, 2021
def date_format(date):
    date = date.split('-')
    if date[1] == '01':
        date[1] = 'January'
    elif date[1] =='02':
        date[1] = 'February'
    elif date[1] =='03':
        date[1] = 'March'
    elif date[1] =='04':
        date[1] = 'April'
    elif date[1] =='05':
        date[1] = 'May'
    elif date[1] =='06':
        date[1] = 'June'
    elif date[1] =='07':
        date[1] = 'July'
    elif date[1] =='08':
        date[1] = 'August'
    elif date[1] =='09':
        date[1] = 'September'
    elif date[1] =='10':
        date[1] = 'October'
    elif date[1] =='11':
        date[1] = 'November'
    elif date[1] =='12':
        date[1] = 'December'
    if date[2] == '01':
        date[2] = '1'
    elif date[2] == '02':
        date[2] = '2'
    elif date[2] == '03':
        date[2] = '3'
    elif date[2] == '04':
        date[2] = '4'
    elif date[2] == '05':
        date[2] = '5'
    elif date[2] == '06':
        date[2] = '6'
    elif date[2] == '07':
        date[2] = '7'
    elif date[2] == '08':
        date[2] = '8'
    elif date[2] == '09':
        date[2] = '9'
    return ('{} {}, {}'.format(date[1],date[2], date[0]))

#name formatter, takes in YEUNG, MATTHEW; WHITE, MICHAEL THOMAS; -> return Matthew Yeung & Michael Thomas White
def name_format(name):
    result = ('')
    temp_list=[]
    #if more than 1 name, use a list to separate names, else create a list with 1 element
    if ';' in name:
        name = name.split(';')
    else:
        temp_list.append(name)
        name = temp_list 
    # remove the empty element created by last ;
    if len(name) > 1:
        name.pop()
    i_count = 0
    for i in name:
        if i_count >= 1:
            i = i.strip()
        # if in , format, swap last anf first name
        if ',' in i:
            exchange_name = i.split(',')
            exchange_name = exchange_name[1].strip() + ' ' + exchange_name[0].strip()
        else:
            exchange_name = i
        name[i_count] = exchange_name
        i_count += 1
    for i in name:
        result += i
        result += (' & ')
    result = result[:-3]
    return result.title()

# For generating sales history
def sales_history():
    global sales_history_length
    sales_history_length = 1
    column_no = 'D'
    result = ('')
    # collect geowarehouse data
    geo_date = date_format(ws[column_no+'25'].value)
    geo_name = str(name_format(ws[column_no+'26'].value))
    geo_price = money_format(str(ws[column_no+'27'].value))
    if geo_price == '0':
        result += ('According to GeoWarehouse, the subject property was registered title on {}, to {}, for an undisclosed amount. \n'.format(geo_date, geo_name))
    else:
        result += ('According to GeoWarehouse, the subject property was registered title on {}, to {}, for a total consideration of ${}. \n'.format(geo_date, geo_name, money_format(geo_price)))
    # collect mls data
    mls_count = int(ws[column_no+'28'].value)
    if mls_count > 0:
        #add all MLS listings
        for i in range(mls_count):
            sales_history_length += 1
            if i == 0:
                mls_column_no = 'D'
            elif i == 1:
                mls_column_no = 'E'
            elif i == 2:
                mls_column_no = 'F'
            elif i == 3:
                mls_column_no = 'G'
            elif i == 4:
                mls_column_no = 'H'
            mls_no = str(ws[mls_column_no+'29'].value)
            mls_date = date_format(ws[mls_column_no+'30'].value)
            listing_price = money_format(str(ws[mls_column_no+'31'].value))
            sold_price = str(ws[mls_column_no+'32'].value)
            dom = ws[mls_column_no+'33'].value
            # check if mls is terminated/cancelled etc
            if sold_price == 'cond':
                result += ('As per MLS#{}, the subject property was listed on {} with an asking price of ${} and later was sold conditionally. \n'.format(mls_no, mls_date, listing_price))
            elif sold_price == 'ter':
                 result += ('As per MLS#{}, the subject property was listed on {} with an asking price of ${} and later was terminated after being on the market for {} days. \n'.format(mls_no, mls_date, listing_price, dom))
            elif sold_price == 'can':
                    result += ('As per MLS#{}, the subject property was listed on {} with an asking price of ${} and later was cancelled after being on the market for {} days. \n'.format(mls_no, mls_date, listing_price, dom))
            elif sold_price == 'sus':
                result += ('As per MLS#{}, the subject property was listed on {} with an asking price of ${} and later was suspended after being on the market for {} days. \n'.format(mls_no, mls_date, listing_price, dom))
            elif sold_price == 'exp':
                result += ('As per MLS#{}, the subject property was listed on {} with an asking price of ${} and later was expired after being on the market for {} days. \n'.format(mls_no, mls_date, listing_price, dom))           
            elif sold_price == 'act':
                result += ('As per MLS#{}, the subject property was listed on {} with an asking price of ${} and is currently listed for sale. \n'.format(mls_no, mls_date, listing_price))  
            else:
                result += ('As per MLS#{}, the subject property was listed on {} with an asking price of ${} and later was sold for ${} after being on the market for {} days. \n'.format(mls_no, mls_date, listing_price, money_format(sold_price), dom))
    elif mls_count == 0:
        result += ('Currently, there is no MLS listing for this property. \n')
    #collect psa data
    psa = ws[column_no+'34'].value
    if psa == 'y':   
        sales_history_length += 2
        psa_date = date_format(ws[column_no+'35'].value)
        psa_buyer = ws[column_no+'36'].value
        psa_seller = ws[column_no+'37'].value
        psa_price = money_format(ws[column_no+'38'].value)
        result += ('As per the purchase and sales agreement dated {}, the subject property is under contract between {} (Buyer) and {} (Seller) for a total consideration of ${}.'.format(psa_date, psa_buyer, psa_seller, psa_price)) 
    result += ('\n\n')
    return result
 
# for taking subject data from excel sheet
def input_excel():
    global ws, city, district, ownership, type, condo_location, occupy, adverse, facilities, townhouse_location, adverse_range, nuclear_station, street_type, storey, electric, parking_type, parking_no, lease, basement, freehold_location, ownership_restriction_checkbox
    global year_built, sqft
    wb = load_workbook(filename = 'house_info.xlsx')
    ws = wb['house_info']
    column_no= 'D'
    city = ws[column_no+'2'].value
    district = ws[column_no+'3'].value
    ownership = str(ws[column_no+'4'].value)
    type = str(ws[column_no+'5'].value)
    storey = str(ws[column_no+'6'].value)
    freehold_location = ws[column_no+'7'].value.split(',') 
    condo_location = ws[column_no+'8'].value.split(',') 
    townhouse_location = ws[column_no+'9'].value
    street_type = str(ws[column_no+'10'].value)
    occupy = str(ws[column_no+'11'].value)
    parking_type = str(ws[column_no+'12'].value)
    parking_no = str(ws[column_no+'13'].value)
    basement = str(ws[column_no+'14'].value)
    adverse = ws[column_no+'15'].value
    nuclear_station = str(ws['F15'].value)
    adverse_range = ws[column_no+'16'].value.split(',') 
    facilities = ws[column_no+'17'].value
    electric = str(ws[column_no+'18'] .value)
    lease = ws[column_no+'19'].value
    year_built = str(ws[column_no+'20'].value)
    sqft = str(ws[column_no+'21'].value)
    ownership_restriction_checkbox = ws[column_no+'22'].value

    #transform some inputs
    if street_type == '0':
        street_type = 'commercial street'
    elif street_type == '1':
        street_type = 'residential street'
    elif street_type == '2':
        street_type = 'main road'
    elif street_type == '3':
        street_type = 'commuter street'
    if electric == '0':
        electric = 'underground'
    else:
        electric = 'overhead'
    if adverse == None:
        adverse = ['']
    else:
        adverse = adverse.split(',') 

    #always 1 storey for apartment & no basement
    if type == '3':
        storey = '1'
        basement = '3'

# generate basement comments
def basement_comment_gen():
    global basement_kitchen, all_basement_int
    column_no = 'D'
    basement_kitchen = 0 
    if basement == '3':
        return ('None; Interior not viewed by the appraiser at the request of the lender due to safety and health concerns impacted by the Co-Vid 19 Virus. It is unknown if there was water seepage at the time of appraisal. Appraisal is based on a sound foundation. \n\n')
    elif basement == '2':
        return ('The basement is unfinished consists of a mech area.  Interior not viewed by the appraiser at the request of the lender due to safety and health concerns impacted by the Co-Vid 19 Virus. '
        'It is unknown if there was water seepage at the time of appraisal. Appraisal is based on a sound foundation. \n\n')
    else:
        basement_interior = []
        basement_rooms = ws[column_no+'41'].value
        basement_interior.extend(['separate entrance', basement_rooms])
        basement_rooms = ws[column_no+'42'].value.split(',')
        if basement_rooms[2] == '1':
            basement_kitchen = 1
        basement_interior.extend(['living room', basement_rooms[0] , 'dining room', basement_rooms[1], 'kitchen', basement_rooms[2], 'recreational room', basement_rooms[3]])
        basement_rooms = ws[column_no+'43'].value.split(',')
        basement_interior.extend(['family room', basement_rooms[0], 'den', basement_rooms[1], 'laundry room', basement_rooms[2], 'storage room', basement_rooms[3]])
        basement_rooms = ws[column_no+'44'].value.split(',')
        basement_interior.extend(['bedroom', basement_rooms[0] , 'full bathroom', basement_rooms[1], '2-piece bathroom', basement_rooms[2]])
        all_basement_int = basement_interior
        # remove the room from list if 0 > no that type of room
        new_interior = []
        i = 0
        while i < len(basement_interior):
            try:
                int(basement_interior[i])
                if basement_interior[i] != '0':
                    new_interior.extend([basement_interior[i-1], basement_interior[i]])
                i += 1
            except:
                i += 1
        basement_interior = new_interior
        types_of_room = int(len(basement_interior)/2)
        i = 0
        result = ''
        suffix = 'a '
        while i < types_of_room:
            result += suffix
            if int(basement_interior[i*2+1]) == 1 and basement_interior[i*2] == '2-piece bathroom':
                result += 'a ' + basement_interior[i*2]
            elif int(basement_interior[i*2+1]) > 1 or basement_interior[i*2] == 'bedroom' or basement_interior[i*2] == 'full bathroom':
                result += basement_interior[i*2+1] + ' ' + basement_interior[i*2]
            else:
                result += basement_interior[i*2]
            suffix = ''
            if int(basement_interior[i*2+1]) > 1:
                suffix += ('s')
            suffix += (', ')
            i += 1
        result += suffix
        if basement == '0':
            result = ('The basement is fully finished featuring {}and mech area. Interior not viewed by the appraiser at the request of the lender due to safety and health concerns impacted by the Co-Vid 19 Virus. '
            'It is unknown if there was water seepage at the time of appraisal. Appraisal is based on a sound foundation. '.format(result))
        elif basement == '1':
            result = ('The basement is partly finished featuring {}and mech area. Interior not viewed by the appraiser at the request of the lender due to safety and health concerns impacted by the Co-Vid 19 Virus. '
            'It is unknown if there was water seepage at the time of appraisal. Appraisal is based on a sound foundation. '.format(result))
        result += ('\n\n')
        return result

# generate interior comments
def interior_info():
    global interior1, interior2, interior3, total_bed, total_partial, total_bath, interior_finishes, exterior_finish, interior2_floor, interior1_floor, interior3_floor
    result = ('')
    column_no = 'D'
    exterior_finish = ws[column_no+'47'].value.split(',')
    #for 1 storey
    interior1 = ['foyer', '1']
    storey_rooms = str(ws[column_no+'48'].value)
    if storey_rooms == '1':
        interior1.extend(['living room', '1', 'dining room', '1', 'kitchen', '1'])
    else:
        interior1.extend(['living room', '0', 'dining room', '0', 'kitchen', '0'])
    storey_rooms = ws[column_no+'49'].value.split(',')
    if ownership == '1' and type == '3':
        interior1.extend(['family room', storey_rooms[0] , 'den' , storey_rooms[1] , 'office' , storey_rooms[2] , 'ensuite washer/dryer' , storey_rooms[3]])
    else:
        interior1.extend(['family room', storey_rooms[0] , 'den' , storey_rooms[1] , 'office' , storey_rooms[2] , 'laundry room' , storey_rooms[3]])
    storey_rooms = ws[column_no+'50'].value.split(',')
    total_bed = int(storey_rooms[0]) + int(storey_rooms[1])
    interior1.extend(['master bedroom', storey_rooms[0] , 'bedroom' , storey_rooms[1]])
    storey_rooms = ws[column_no+'51'].value.split(',')
    total_bath = int(storey_rooms[0]) + int(storey_rooms[1])
    total_partial = int(storey_rooms[2])
    interior1.extend(['full ensuite bathroom', storey_rooms[0] , 'full bathroom' , storey_rooms[1], '2-piece bathroom', storey_rooms[2]])
    interior1_floor = ws[column_no+'52'].value.split(',')
    # for 2 storey
    if storey != '1':
        interior2 = []
        storey_rooms = str(ws[column_no+'53'].value)
        if storey_rooms == '1':
            interior2.extend(['living room', '1', 'dining room', '1', 'kitchen', '1'])
        else:
            interior2.extend(['living room', '0', 'dining room', '0', 'kitchen', '0'])
        storey_rooms = ws[column_no+'54'].value.split(',')
        if ownership == '1' and type == '3':
            interior2.extend(['loft', storey_rooms[0] , 'den' , storey_rooms[1] , 'office' , storey_rooms[2] , 'ensuite washer/dryer' , storey_rooms[3]])
        else:
            interior2.extend(['loft', storey_rooms[0] , 'den' , storey_rooms[1] , 'office' , storey_rooms[2] , 'laundry room' , storey_rooms[3]])
        storey_rooms = ws[column_no+'55'].value.split(',')
        total_bed += int(storey_rooms[0]) + int(storey_rooms[1])
        interior2.extend(['master bedroom', storey_rooms[0] , 'bedroom' , storey_rooms[1]])
        storey_rooms = ws[column_no+'56'].value.split(',')
        total_bath += int(storey_rooms[0]) + int(storey_rooms[1])
        total_partial += int(storey_rooms[2])
        interior2.extend(['full ensuite bathroom', storey_rooms[0] , 'full bathroom' , storey_rooms[1], '2-piece bathroom', storey_rooms[2]])
        interior2_floor = ws[column_no+'57'].value.split(',')
    # third floor
    if storey == '2 1/2 ' or storey == '3':
        interior3=[]
        storey_rooms = ws[column_no+'58'].value.split(',')
        interior3.extend(['loft', storey_rooms[0] , 'den' , storey_rooms[1] , 'office' , storey_rooms[2] , 'laundry room' , storey_rooms[3]])
        storey_rooms = ws[column_no+'59'].value.split(',')
        total_bed += int(storey_rooms[0]) + int(storey_rooms[1])
        interior3.extend(['master bedroom', storey_rooms[0] , 'bedroom' , storey_rooms[1]])
        storey_rooms = ws[column_no+'60'].value.split(',')
        total_bath += int(storey_rooms[0]) + int(storey_rooms[1])
        total_partial += int(storey_rooms[2])
        interior3.extend(['full ensuite bathroom', storey_rooms[0] , 'full bathroom' , storey_rooms[1], '2-piece bathroom', storey_rooms[2]])
        interior3_floor = ws[column_no+'61'].value.split(',')

    #interior finishes
    interior_finishes = str(ws[column_no+'62'].value)

    #interior comment freehold
    # turn house type Det=0/Semi=1/Town=2 to names
    if type == '0':
        house_name = 'detached'
    elif type == '1':
        house_name = 'semi-detached'
    elif type == '2':
        house_name = 'townhouse'
    if ownership == '0':
        result += ('The subject is a {} storey {} dwelling containing approximately {} square feet of living space and is assumed to be in average physical condition at time of inspection. The exterior of the home is finished with mostly {}. '
        'The main floor, finished with {} flooring consists of {}'.format(storey, house_name, money_format(sqft), exterior_comment_gen(exterior_finish), floor_comment_gen(interior1_floor), interior_room_gen(interior1)))
        if storey != '1':
            result += ('The second floor, finished with {} flooring, consists of {}'.format(floor_comment_gen(interior2_floor), interior_room_gen(interior2)))
        if storey == '3' or storey == '2 1/2':
            result += ('The third floor, finished with {} flooring, consists of {}'.format(floor_comment_gen(interior3_floor), interior_room_gen(interior3)))
        # interior finishes comment
        result += interior_finishes_gen(interior_finishes)
        result += ('Other interior finishes are standard and consistent with the similar type properties in the subject area. The effective age utilized in this appraisal report is based on the MLS listing and pictures provided by the homeowner. '
        'We have assumed the subject property being at least average condition and the norm for the area with other similar properties. ')
    elif ownership == '1' and type == '2':
        result += ('The subject is a {} storey condominium townhouse {} unit containing approximately {} square feet of living space and is assumed to be in average physical condition at time of appraisal. '
        'The exterior of the home is finished with mostly {}. The main floor, finished with {}, consists of {}'. format(storey, condo_location[3], money_format(sqft), exterior_comment_gen(exterior_finish), floor_comment_gen(interior1_floor), interior_room_gen(interior1)))
        if storey != '1':
            result += ('The second floor, finished with {} flooring, consists of {}'.format(floor_comment_gen(interior2_floor), interior_room_gen(interior2)))
        if storey == '3' or storey == '2 1/2':
            result += ('The third floor, finished with {} flooring, consists of {}'.format(floor_comment_gen(interior3_floor), interior_room_gen(interior3)))
        result += interior_finishes_gen(interior_finishes)
        result += ('Other interior finishes are standard and consistent with similar type units in the condominium complex. The effective age utilized in this appraisal report is based on the MLS listing and pictures provided by the homeowner. '
        'We have assumed the subject property being at least average condition and the norm for the area with other similar properties. ')
    elif ownership == '1' and type == '4':
        result += ('The subject is a {} storey condominium stacked townhouse {} unit containing approximately {} square feet of living space and is assumed to be in average physical condition at time of appraisal. '
        'The exterior of the home is finished with mostly {}. The main floor, finished with {}, consists of {}'. format(storey, condo_location[3], money_format(sqft), exterior_comment_gen(exterior_finish), floor_comment_gen(interior1_floor), interior_room_gen(interior1)))
        if storey != '1':
            result += ('The second floor, finished with {} flooring, consists of {}'.format(floor_comment_gen(interior2_floor), interior_room_gen(interior2)))
        if storey == '3' or storey == '2 1/2':
            result += ('The third floor, finished with {} flooring, consists of {}'.format(floor_comment_gen(interior3_floor), interior_room_gen(interior3)))
        result += interior_finishes_gen(interior_finishes)
        result += ('Other interior finishes are standard and consistent with similar type units in the condominium complex. The effective age utilized in this appraisal report is based on the MLS listing and pictures provided by the homeowner. '
        'We have assumed the subject property being at least average condition and the norm for the area with other similar properties. ')
    elif ownership == '1' and type == '3':
        result += ('The subject is a {} bedroom condominium apartment {} unit containing approximately {} square feet of living space and is assumed to be in average physical condition at time of appraisal. '
        'The floor plan consists of {}Interior is finished with laminate flooring in the living room, dining room and bedrooms, ceramic tiles in the kitchen, bathroom and foyer. '
        'Interior finishes are standard and consistent with similar type units in the condominium complex. '. format(total_bed, condo_location[3], money_format(sqft), interior_room_gen(interior1)))
        result += interior_finishes_gen(interior_finishes)
        result += ('The effective age utilized in this appraisal report is based on the MLS listing and pictures provided by the homeowner. We have assumed the subject property being at least average condition and the norm for the area with other similar properties. ')
    result += ('\n\n')
    return result

#generate interior finishes comment
def interior_finishes_gen(interior_finishes):
    if interior_finishes == '1':
        return ('Interior upgrades include crown mouldings. ')
    elif interior_finishes == '0':
        return ('Interior upgrades include potlights. ')
    elif interior_finishes == '0,1': 
        return ('Interior upgrades include potlights and crown mouldings. ')
    else:
        return ('')

#generate comment for flooring (takes list ['0','1','2'] etc)
#hardwood=0/broadloom=1/laminate=2/ceramic=3
def floor_comment_gen(flooring):
    result = ('')
    floor_type_count = 0
    floor_type_total = len(flooring)
    for i in flooring:
        if i == '0':
            result += ('hardwood')
        elif i == '2':
            result += ('laminate')
        elif i == '1':
            result += ('broadloom')
        elif i == '3':
            result += ('ceramic tiles')
        if floor_type_count < (floor_type_total-2):
            result += (', ')
        elif floor_type_count == (floor_type_total-2):
            result += (', and ')
        floor_type_count += 1
    return result

#for generating exterior finish comment
# brick=0/stone=1/vinyl=2/concrete=3/stucco=4
def exterior_comment_gen(exterior_finish):
    result = ('')
    exterior_type_count = 0
    exterior_type_total = len(exterior_finish)
    for i in exterior_finish:
        if i == '0':
            result += ('brick')
        elif i == '1':
            result += ('stone')
        elif i == '2':
            result += ('vinyl siding')
        elif i == '3':
            result += ('concrete')
        elif i == '4':
            result += ('stucco')
        if exterior_type_count < (exterior_type_total-2):
            result += (', ')
        elif exterior_type_count == (exterior_type_total-2):
            result += (' and ')
        exterior_type_count += 1
    return result

# takes interior1/2/3, a list where odd index is room name and even index is room count, generate 'a foyer, living room, dining room, a master bedroom featuring a full ensuite bathroom, 3 average sized bedrooms and a full bathroom off the main hallway' etc
#['family room', '1', 'den', '0', 'office', '0', 'laundry room', '1', 'master bedroom', '0', 'bedroom', '1', 'full ensuite bathroom', '0', 'full bathroom', '0', '2-piece bathroom', '1'] 
# ['loft', '0', 'den', '1', 'office', '1', 'laundry room', '0', 'master bedroom', '1', 'bedroom', '2', 'full ensuite bathroom', '1', 'full bathroom', '1', '2-piece bathroom', '0']
def interior_room_gen(interior):
    result = ''
    # remove the room from list if 0 > no that type of room
    new_interior = []
    i = 0
    while i < len(interior):
        try:
            int(interior[i])
            if interior[i] != '0':
                new_interior.extend([interior[i-1], interior[i]])
            i += 1
        except:
            i += 1
    interior = new_interior
    types_of_room = int(len(interior)/2)
    i = 0
    suffix = 'a '
    while i < types_of_room:
        # only print comment if there is the room
        if interior[i*2] != 'full ensuite bathroom':
            # master bedroom with ensuite comments
            if interior[i*2] == 'master bedroom' :
                try:
                    if interior[i*2+4] == 'full ensuite bathroom' and int(interior[i*2+5]) > 0:
                        result += (suffix + 'master bedroom featuring a full ensuite bathroom')
                        interior[i*2+1] = str(int(interior[i*2+1])-1)
                        # remove ensuite count by 1, rest for average bedroom
                        interior[i*2+5] = str(int(interior[i*2+5])-1)
                except IndexError:
                    pass
                try:
                    if interior[i*2+2] == 'full ensuite bathroom' and int(interior[i*2+3]) > 0:
                        result += (suffix + 'master bedroom featuring a full ensuite bathroom')
                        interior[i*2+1] = str(int(interior[i*2+1])-1)
                        # remove ensuite count by 1, rest for average bedroom
                        interior[i*2+5] = str(int(interior[i*2+3])-1)
                except IndexError:
                    pass
                if interior[i*2+1] == '1':
                    result += (suffix + 'master bedroom')
                    interior[i*2+1] = str(int(interior[i*2+1])-1)
            #average bedroom with ensuite comments
            else:
                try:
                    if interior[i*2] == 'bedroom' and interior[i*2+2] == 'full ensuite bathroom' and int(interior[i*2+3]) == 1:
                        result += (suffix + interior[i*2+3] + ' averaged sized bedroom featuring a full ensuite bathroom')
                        interior[i*2+1] = str(int(interior[i*2+1])-1)
                    elif interior[i*2] == 'bedroom' and interior[i*2+2] == 'full ensuite bathroom' and int(interior[i*2+3]) > 1:
                        result += (suffix + interior[i*2+3] + ' averaged sized bedrooms featuring full ensuite bathrooms')
                        interior[i*2+1] = str(int(interior[i*2+1])-int(interior[i*2+3]))
                except IndexError:
                    pass
            #reformat bedroom name
            if interior[i*2] == 'bedroom':
                interior[i*2] = 'average sized bedroom'
            # adding all rooms, omitting ensuite bathroom
            if  int(interior[i*2+1]) > 0 :
                if interior[i*2] == '2-piece bathroom' and int(interior[i*2+1]) == 1:
                    result += (suffix + 'a ' + interior[i*2])
                elif interior[i*2] == 'average sized bedroom' or interior[i*2] == 'full bathroom' or int(interior[i*2+1]) > 1:
                    result += (suffix + interior[i*2+1] + ' ' + interior[i*2])
                else:
                    result += (suffix + interior[i*2])
                #formating, and s for plural, and add ','
            suffix = ''
            if int(interior[i*2+1]) > 1:
                suffix += ('s')
            if interior[i*2] == 'full bathroom' and (storey == '1 1/2' or storey == '2' or storey == '2 1/2' or storey == '3'):
                suffix += (' off the main hallway. ')
            elif i == types_of_room-2:
                suffix += (', and ')
            elif i == types_of_room-1:
                suffix += ('. ')
            else:
                suffix += (', ')
        # special case for ensuite baths
        elif interior[i*2] == 'full ensuite bathroom':
            suffix = ''
            if int(interior[i*2-1]) > 1:
                suffix += ('s')
            if i == types_of_room-2:
                suffix += ', and '
            elif i == types_of_room-1:
                suffix += ('. ')
            else:
                suffix += (', ')
        
        i += 1
    result += suffix
    return result

# for neighbourhood and site comments
def neighbour_site():
    result = ('')
    # check demand
    if type == '3':
        demand = 'stable'
    else:
        demand ='increasing'
    #condo neighourhood
    if ownership == '1':
        result += ("Neighbourhood\nThe subject area is located within the {} quadrant of {} and {}, in the MLS district known as '{}' in the {}. The neighbourhood is comprised of a mix of residential, "
        'commercial, as well as condominium properties. The immediate area consists mainly of residential established properties ranging up to _ years old in varying age, size, and condition. Newly built condominium buildings are also noted in the area. '
        'Based on the average days on market on the recent MLS listings, the demands for properties in this neighborhood is considered average to good. Based on research on recent MLS listings, '
        'the market trend in the immediate area was {} at the time of appraisal. '
        "The subject is in close proximity to local area shopping centres, {}, public transit, public schools, and parks. ".format(condo_location[0], condo_location[1], condo_location[2], district, neighbourhood_city(), demand, facilities_format()))
    else:
        #freehold neighbourhood
        result += ("Neighbourhood\nThe subject area is located within the MLS district known as '{}' in the {}. The neighbourhood is comprised of a mix of residential, commercial, as well as condominium properties. "
        'The immediate area consists mainly of residential established properties ranging up to _ years old in varying age, size and condition. Newer built condominium developments were also noted in the general area. '
        'Based on average days on market on the recent MLS listings, the demand for properties in this neighborhood is considered average to good. Based on research on recent MLS listings, the market trend in the immediate area is considered increasing at the time of appraisal. '
        'The subject property is located in close proximity to local area shopping, places of worship, public parks, schools, public transit{}. '.format(district, neighbourhood_city(), facilities_format()))
    adverse_comment = adverse_comment_gen()
    result += adverse_comment + ('\n\nSite\n')
    #condo townhouse site
    if ownership == '1' and type == '2':
        result += ('Appraiser did not note any wind turbines at the property at the time of appraisal. Underground storage utility were not visible to the appraiser on the subject site at time of appraisal viewing. '
        "Environmental hazards were not noted at the property based on appraiser's visual observation at the time of appraisal viewing. "
        'The subject site is improved with a condominium townhouse complex located on a {}. The street characteristics include municipal servicing, paved roads with curbs, street lights, and hydro wires placed {}. '
        'The subject property is a {} storey condominium townhouse {} unit{}'
        'As a title search was not conducted, we are not aware of any easements or restrictions that would have a negative impact on value. '
        'We were not provided any documentation regarding environmental contamination. The complex is subject to condo by-laws, lenders should satisfy themselves. '.format(street_type, electric, storey, condo_location[3], garage_comment_gen()))
    #condo apt
    elif ownership == '1' and type =='3':
        result += ('Appraiser did not note any wind turbines at the property at the time of appraisal. Underground storage utility were not visible to the appraiser on the subject site at time of appraisal viewing. '
        "Environmental hazards were not noted at the property based on appraiser's visual observation at the time of appraisal viewing. The subject site is improved with a high rise condominium complex located on a {}. "
        'The street characteristics include municipal servicing, paved roads with curbs, street lights, and hydro wires placed {}. '
        'The subject property is a {} bedroom condominium apartment {} unit{}'
        'As a title search was not conducted, we are not aware of any easements or restrictions that would have a negative impact on value. We were not provided any documentation regarding environmental contamination. '
        'The complex is subject to condo by-laws, lenders should satisfy themselves. Common site area includes fitness centre, swimming pool, guest suites and party room. '.format(street_type, electric, str(total_bed), condo_location[3], garage_comment_gen()))
    #stack townhouse
    elif ownership == '1' and type == '4':
        result += ('Appraiser did not note any wind turbines at the property at the time of appraisal. Underground storage utility were not visible to the appraiser on the subject site at time of appraisal viewing. '
        "Environmental hazards were not noted at the property based on appraiser's visual observation at the time of appraisal viewing. "
        'The subject site is improved with a condominium townhouse complex located on a {}. The street characteristics include municipal servicing, paved roads with curbs, street lights, and hydro wires placed {}. '
        'The subject property is a {} storey condominium stacked townhouse {} unit{}'
        'As a title search was not conducted, we are not aware of any easements or restrictions that would have a negative impact on value. '
        'We were not provided any documentation regarding environmental contamination. The complex is subject to condo by-laws, lenders should satisfy themselves. '.format(street_type, electric, storey, condo_location[3], garage_comment_gen()))
    #freehold
    elif ownership == '0':
        # naming the type of house
        if type == '0':
            freehold_type = 'detached'
        elif type == '1':
            freehold_type = 'semi-detached'
        elif type == '2':
            freehold_type = 'townhouse'
        result += ('Appraiser did not note any wind turbines at the property at the time of appraisal. Underground oil storage utility were not visible to the appraiser at time of appraisal. '
        "Environmental hazards were not noted at the property based on appraiser's visual observation at the time of appraisal viewing. Site is a {} shaped {} lot located on a {}, backing onto {}. "
        'The lot is improved with a {} storey {} dwelling{}The street characteristics include municipal servicing, paved roads with curbs, street lights, and hydro wires placed {}. '
        'Site improvements include covered concrete porch, pathways and a patio at rear. The subject site features fully fenced backyard and average and similar landscaping in comparison to other properties in the area. '
        'The subject property existing use is residential single family '.format(freehold_location[0], freehold_location[1], street_type, freehold_location[2], storey, freehold_type, garage_comment_gen(), electric,))
        if basement_kitchen == 1:
            result+='with self contained basement unit '
        result+='and it is the opinion of the appraiser that this activity constitutes the highest and best use. '
    result += adverse_comment

    #basement kitchen comment
    if basement_kitchen == 1:
        if occupy == '1' or occupy == '2':
            result += ('As per the homeowner, the basement unit is tenant occupied. '
            'The subject property has a basement kitchen of which retrofit or secondary living status is not warranted by the appraiser and should be verified by the lender. '
            'As per MPAC, the subject property is registered as a single family dwelling. At the request of the client, the property is being valued in Fee Simple')
        else:
            result += ('As per the homeowner, the basement is being used for personal use only. '
            'The subject property has a basement kitchen of which retrofit or secondary living status is not warranted by the appraiser and should be verified by the lender. '
            'As per MPAC, the subject property is registered as a single family dwelling. At the request of the client, the property is being valued in Fee Simple')
    #tenant comment
    elif occupy == '1' or occupy == '2':
        if ownership == '0':
            ownership_name = 'Fee Simple'
        else:
            ownership_name = 'Condominium Ownership'
        result += ('As per the homeowner, the subject property was tenant occupied at time of appraisal. At the request of the client, the property is being valued in ' + ownership_name) 
    if basement_kitchen == 1 or occupy == '1' or occupy == '2':
        if lease == 'n':
            result += (', and the Leased Fee value has not been addressed in this assignment.')
        else:
            result += ('. ')
    return result + ('\n ')

# generate adverse comment
def adverse_comment_gen():
    result = ('')
    # for apartment adverse
    no_adverse = 1
    if type == '3':
        for i in adverse:
            if i == '1':
                result += ('Railroad tracks were also noted within the general area of the subject property. ')
        adverse_no = 0
        for i in adverse:
            if i == '0':
                if city.title() == 'Ottawa':
                    result += ('Adversely, the subject property is situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of Ottawa International Airport and is subject to higher air traffic noises during operational hours. ')
                else:
                    result +=('Adversely, the subject property is situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of Toronto Pearson Airport and is subject to higher air traffic noises during operational hours. ')
                no_adverse = 0
            if i == '4':
                if no_adverse == 1:
                    result += ('Adversely, the subject property issituated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of hydro lanes. ')
                    no_adverse = 0
                else:
                    result += ('The subject property is also situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of hydro lanes. ')
            if i == '5':
                if no_adverse == 1:
                    if nuclear_station == '0':
                        result += ('Adversely, the subject property is situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of Pickering Nuclear Generating Station. ')
                    else :
                        result += ('Adversely, the subject property is situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of Darlington Nuclear Generating Station. ')
                    no_adverse = 0
                else:
                    if nuclear_station == '0':
                        result += ('The subject property is also situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of Pickering Nuclear Generating Station. ')
                    else :
                        result += ('The subject property is also situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of Darlington Nuclear Generating Station. ')                   
            adverse_no += 1
        if no_adverse == 1:
            result += ('The Appraiser did not notice any functional or external obsolescence at the time of the site visit. ')
        else:
            result += ('See plot map. ')
    else :
    # non apt adverse
        adverse_no = 0
        # for out of range railway, no adverse
        if len(adverse) == 1 and adverse[0] == '1' and (adverse_range[0] == '1000' or adverse_range[0] == '1'):
            result += ('Railroad tracks were also noted within the general area of the subject property. ')
        if len(adverse) == 1 and (adverse[0] not in ['0','1','2','3','4','5'] or adverse[0] == '1' and (adverse_range[0] == '1000' or adverse_range[0] == '1')):
            result += ('The Appraiser did not notice any functional or external obsolescence at the time of the site visit. ')
        else:
            no_of_adverse = len(adverse)
            adverse_no = 0
            #with adverse and out of range railways
            for i in adverse:
                if i == '1' and (adverse_range[adverse_no] == '1000' or adverse_range[adverse_no] == '1'):
                    result += ('Railroad tracks were also noted within the general area of the subject property. ')
                    break
                adverse_no += 1
            adverse_no = 0
            result += ('Adversely, the subject property is ')
            for i in adverse:
                # if more than 1 adverse
                if no_of_adverse < len(adverse) and not (i == '1' and (adverse_range[adverse_no] == '1000' or adverse_range[adverse_no] == '1')):
                    result += ('The subject property is also ')
                no_of_adverse -=1
                # add adverse comments
                if i == '0':
                    if city.title() == 'Ottawa':
                        result += ('situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of Ottawa International Airport and is subject to higher air traffic noises during operational hours. ')
                    else:
                        result +=('situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of Toronto Pearson Airport and is subject to higher air traffic noises during operational hours. ')
                elif i == '1':
                    if adverse_range[adverse_no] != '1000' and adverse_range[adverse_no] != '1':
                        result += ('situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of railroad tracks and is subject to railroad noises during operational hours. ')
                    else:
                        no_of_adverse +=1
                elif i == '2':
                    result += ('located on a main road, which is subjected to busy and noisy traffic during peak hours. ')
                elif i == '3':
                    result += ('located on a commuter street, which is subjected to busy and noisy traffic during peak hours. ')
                elif i == '4':
                    result += ('situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of hydro lanes. ')
                elif i == '5':
                    if nuclear_station == '0':
                        result += ('situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of Pickering Nuclear Generating Station. ')
                    else :
                        result += ('situated within ' + adverse_distance_convert(adverse_range[adverse_no]) + ' of Darlington Nuclear Generating Station. ')
                if i!='2' and i!='3':
                    adverse_no +=1
            result += ('See plot map. ')
    return result

#convert garage id to text, for site comment
# Attach=0/Builtin=1/Detach=2/Underg=3/Aboveg=4/None+Driveway=5/Nonenone = 6
def garage_comment_gen():
    result = ('')
    if parking_type == '3' or parking_type =='4':
        result += (' that includes ')
    elif parking_type == '0' or parking_type == '1' or parking_type == '2' or parking_type =='5':
        result += (' featuring a ')
    if int(parking_no) > 0:
        result += parking_no
    if parking_type == '0':
        result += (' car attached ')
    elif parking_type == '1':
        result += (' car built-in ')
    elif parking_type == '2':
        result += (' car detached ')
    elif parking_type == '3':
        result += (' underground parking space and a locker***. ')
    elif parking_type == '4':
        result += (' surface parking space. ')
    elif parking_type == '5':
        result += ('private asphalt-surfaced driveway. ')
    else :
        result += ('. ')
    if parking_type == '0' or parking_type == '1' or parking_type == '2':
        result += ('garage on a private asphalt-surfaced driveway. ')
    return result

# for adverse_comment, add M/KM/close proxmity comment
def adverse_distance_convert(dist):
    if dist == '200':
        return ('close proximity')
    elif dist == '5':
        return ('5 KM')
    elif dist == '10':
        return ('10 KM')
    elif dist == '500':
        return ('500 M')
    elif dist == '1000':
        return ('1 KM')
    elif dist == '1':
        return ('1 KM')

def facilities_format():
    if ownership == '0':
        if ',' in facilities:
            x = facilities.split(',')
            i = 0
            result = ''
            while i < len(x):
                if i == 0:
                    result += (', ') + x[i]
                elif i != len(x)-1:
                    result += (',') + x[i]
                else:
                    result += (' and') + x[i]
                i += 1
            return result
        else:
            return (' and ') + facilities
    else:
        return facilities
# generate city for neighbourhood comment
def neighbourhood_city():
    result = ""
    if city.title() == "Ajax" or city.title() == "Whitby":
        result = "Town of " + city.title()
    elif city.title() == "Clarington":
        result = "Municipality of " + city.title()
    else:
        result = "City of " + city.title()
    return result
    
# generate city and region for page3
def city_comment_gen():
    result = ""
    if city.title() == "Ottawa":
        result = "City of Ottawa;?"
    elif city.title() == "Toronto":
        result = "City of Toronto;?"
    elif city.title() == "Brampton":
        result = "Region of Peel; City of Brampton"
    elif city.title() == "Mississauga":
        result = "Region of Peel; City of Mississauga"
    elif city.title() == "Pickering":
        result = "Region of Durham; City of Pickering"
    elif city.title() == "Oshawa":
        result = "Region of Durham; City of Oshawa"
    elif city.title() == "Whitby":
        result = "Region of Durham; Town of Whitby"
    elif city.title() == "Ajax":
        result = "Region of Durham; Town of Ajax"
    elif city.title() == "Markham":
        result = "Region of York; City of Markham"
    elif city.title() == "Vaughan":
        result = "Region of York; City of Vaughan"
    elif city.title() == "Clarington":
        result = "Region of Durham; Municipality of Clarington"
    return result

#----auto input codes--------
# delete everything in box
def delete():
    keyboard.press(Key.ctrl)
    keyboard.type('a')
    keyboard.release(Key.ctrl)
    keyboard.press(Key.delete)
    keyboard.release(Key.delete)

# press tab for (times)
def tab(times):
    for _ in range(times):
        keyboard.press(Key.tab)
        keyboard.release(Key.tab)

# press up for (times) - for box type input
def up(times):
    for _ in range(times):
        keyboard.press(Key.up)
        keyboard.release(Key.up)

# press down for (times) - for box type input
def down(times):
    for _ in range(times):
        keyboard.press(Key.down)
        keyboard.release(Key.down)

# press right for (times)
def right(times):
    for _ in range(times):
        keyboard.press(Key.right)
        keyboard.release(Key.right)

# press right for (times)
def left(times):
    for _ in range(times):
        keyboard.press(Key.left)
        keyboard.release(Key.left)

# press enter single time
def enter():
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)

# for page 4 effective age
def effective_age():
    if year_built == '2022':
        e_age = '0'
    elif year_built >= '2020':
        e_age = '1'
    elif year_built >= '2018':
        e_age = '2'
    elif year_built >= '2016':
        e_age = '3'
    elif year_built >= '2007':
        e_age = '5'
    elif year_built >= '1997':
        e_age = '10'
    elif type != '0' or year_built >= '1987':
        e_age = '15'
    else:
        e_age = '20'
    if type == '0':
        remain_life = str(55 - int(e_age))
    else:
        remain_life = str(50 - int(e_age))
    return e_age + 'E', remain_life + '+-' 

# takes year_built and house type, return depreciation table value
def depreciation_table(e_age):
    e_age = e_age[:-1]
    #for detached
    result = ''
    if type == '0':
        if e_age == '0' or e_age == '1' or e_age == '2':
            result = e_age
        elif e_age == '3':
            result = '2'
        elif e_age == '5':
            result = '4'
        elif e_age == '10':

            result = '9'
        elif e_age == '15':
            result = '15'
        else:
            result = '21'
    # for non detached
    else:
        if e_age == '0' or e_age == '1' or e_age == '2'or e_age == '3':
            result = e_age
        elif e_age == '5':
            result = '5'
        elif e_age == '10':
            result = '11'
        else:
            result = '17'
    return result

# for page 4 all floring comment
def all_flooring_gen():
    total_floor = interior1_floor
    if storey != '1':
        for floor in interior2_floor:
            if floor not in total_floor:
                total_floor.append(floor)
    if storey == '2 1/2' or storey == '3':
        for floor in interior3_floor:
            if floor not in total_floor:
                total_floor.append(floor)
    total_floor.sort(key = int)
    return floor_comment_gen(total_floor).capitalize()

def table_enter(room):
    if int(room)>0:
        keyboard.type(str(room))

# main function for auto input pdf
def auto_input(basement_comments, interior_comments, sales_history_comments):
    global keyboard
    keyboard = Controller()
    print ('prepare to move mouse: 3 secs')
    # seconds of prepare time
    time.sleep(3)
    # start from page3 - APPRAISER:
    if city.title() == 'Ottawa':
        keyboard.type('Sindu R. Rajaruban')
    else:
        keyboard.type('Ruban Kanagenthiran')
    tab(2)
    if city.title() == "Ottawa":
        delete()
        keyboard.type("N/A")
        tab(2)
        keyboard.type("Ottawa")
        tab(1)
        keyboard.type("o")
        enter()
        tab(7)
    else:
        tab(10)
    enter()
    tab(2)
    enter()
    tab(7)
    if ownership == '0':
        down(2)
    else:
        down(6)
    tab(5)
    keyboard.type(city_comment_gen())
    tab(10)
    # Occupied by:
    if occupy == '0':
        down(2)
    elif occupy == '1':
        down(4)
    elif occupy == '2':
        down(7)
    elif occupy == '3':
        down(3)
    # template skips after Taxes$ to EASEMENTS: Utility/Access/Other
    tab(2)
    #ownership restrictions extra checkbox
    if ownership_restriction_checkbox == 'y':
        enter()
        tab(1)
    enter()
    tab(23)
    if electric == 'overhead':
        right(1)
    tab(2)
    ''' 
    Attach=0/Builtin=1/Detach=2/Underg=3/Aboveg=4/None+Driveway=5/NoneNone=6
    parking_type = '0'
    parking_no = '1'
    '''
    # parking starting at DRIVEWAY: PRIVATE
    if parking_type == '0':
        park_comment = parking_no + ' Attach.'
    elif parking_type == '1':
        park_comment = parking_no + ' Builtin.'
    elif parking_type == '2':
        park_comment = parking_no + ' Detach.'
    elif parking_type == '3':
        park_comment = parking_no + ' Underg.'
    elif parking_type == '4':
        park_comment = parking_no + ' Aboveg.'
    elif parking_type == '5' or parking_type == '6':
        park_comment = 'None'
    if ownership == '0':
        if parking_type == '6':
            enter()
        tab(3)
        if parking_no == '1':
            enter()
            tab(1)
            enter()
            tab(4)
        elif parking_type == '6':
            tab(1)
            enter()
            tab(4)
        else:
            tab(5)
        if parking_type == '6':
            enter()
        tab(6)
        if parking_type == '5' or parking_type == '6':
            enter()
        tab(2)
        if parking_type == '6':
            enter()
        tab(2)
        enter()
        tab(1)
        keyboard.type(park_comment)
    #for condo townhouse template
    #Attach=0/Builtin=1/Detach=2/Underg=3/Aboveg=4/None+Driveway=5/NoneNone=6 
    elif ownership == '1' and (type == '2' or type == '4'):
        if parking_type == '0' or parking_type == '1' or parking_type == '2':
            enter()
        tab(1)
        if (parking_type != '3' and parking_type != '4'):
            enter()
        tab(2)     
        #at single
        if (parking_type == '0' or parking_type == '1' or parking_type == '2') and int(parking_no) > 0:
            enter()
            tab(4)
        elif (parking_type == '0' or parking_type == '1' or parking_type == '2') and int(parking_no) > 1:
            tab(1)
            enter()
            tab(3)
        else:
            tab(4)
        if parking_type != '3':
            enter()
        tab(1)
        if parking_type == '5':
            enter()
        tab(6)
        if parking_type == '4' or parking_type == '5' or parking_type == '6':
            enter()
        tab(2)
        if parking_type != '3' and parking_type != '6' and parking_type != '4':
            enter()
        tab(2)
        if parking_type == '5':
            enter()
        tab(1)
        keyboard.type(park_comment)
    elif ownership == '1' and type == '3':
        tab(1)
        if parking_type == '6':
            enter()
        tab(6)
        if parking_type == '6':
            enter()   
        tab(1)
        if parking_type == '6':
            enter()
        tab(6)
        if parking_type == '6':
            enter()
        tab(4)
        enter()
        tab(1)
        keyboard.type(park_comment)
    tab(5)
    right(1)
    tab(1)
    right(2)
    tab(2)
    # start at page 4 -CONSTRUCTION COMPLETE:
    time.sleep(8)
    enter()
    tab(2)
    keyboard.type(year_built)
    tab(1)
    e_age, remain_life = effective_age()
    keyboard.type(e_age) 
    tab(1)
    keyboard.type(remain_life)
    tab(1)
    right(1)
    tab(1)
    keyboard.type(sqft)
    tab(4)
    enter()
    tab(3)
    # start from PROPERTY TYPE:
    if ownership == '0':
        down(2)
        tab(1)
        if type == '0':
            down(2)
        elif type == '1':
            down(3)
        elif type == '2':
            down(4)
        tab(1)
        if storey == '1':
            down(2)
        elif storey == '1 1/2':
            down(4)
        elif storey == '2':
            down(6)
        elif storey == ' 2 1/2':
            down(8)
        elif storey == '3':
            down(9)
        tab(2)
        if city.title() == 'Ottawa':
            down(2)
        if basement == '3':
            down(4)
        tab(1)
        if basement == '0':
            down(3)
        elif basement == '1':
            down(4)
        elif basement == '2':
            down(5)
        tab(4)
    elif ownership == '1' and type == '3':
        down(2)
        tab(9)
    elif ownership == '1' and (type == '2' or type =='4'):
        down(2)
        tab(1)
        down(3)
        tab(1)
        down(3)
        tab(2)
        if basement != '3':
            up(4)
            tab(1)
            up(1)
            if basement == '0':
                down(2)
            elif basement == '1':
                down(3)
            elif basement == '2':
                down(4)
            tab(4)
        else:
            tab(5)

    # stops at CONDITION bug
    time.sleep(6)
    tab(1)
    # at exterior finish: brick=0/stone=1/vinyl=2/concrete=3/stucco=4 (a list)
    if exterior_finish == ['0']:
        down(2)
    elif exterior_finish == ['0', '2']:
        down(4)
    elif exterior_finish == ['2']:
        down(13)
    elif exterior_finish == ['3']:
        down(16)
    tab(3)
    right(1)
    tab(1)
    right(1)
    tab(1)
    keyboard.type('1')
    tab(1)
    if int(total_bed)-1 > 0:
        keyboard.type(str(int(total_bed)-1))
    tab(6)
    if total_partial > 0:
        keyboard.type(str(total_partial))
    tab(2)
    if total_bath > 0:
        keyboard.type(str(total_bath))
    tab(19)
    # ticking interior finish boxes
    if interior_finishes != None:
        if '1' in interior_finishes:
            enter()
        tab(2)
        if '0' in interior_finishes:
            enter()
    else:
        tab(2)
    tab(1)
    # starts at FLOORING
    delete()
    keyboard.type(all_flooring_gen())
    tab(40)
    delete()
    keyboard.type(basement_comments)
    tab(1)
    if parking_no == '0':
        down(5)
    elif parking_no == '1':
        down(1)
    elif parking_no == '2':
        down(2)
    elif parking_no == '3':
        down(3)
    tab(1)
    if parking_type == '5':
        keyboard.type('None; Private asphalt-surfaced driveway')
    elif parking_type == '6':
        keyboard.type('None')
    elif parking_type == '3':
        keyboard.type(parking_no + ' Underground parking space')
    elif parking_type == '4':
        keyboard.type(parking_no + ' Surface parking space')
    else:
        keyboard.type(parking_no + ' car garage; Private asphalt-surfaced driveway')
    tab(1)
    if ownership =='1' and (type =='2' or type =='4'):
        delete()
        keyboard.type("Concrete front porch,  steps,  pathway and patio at rear; ")
    tab(1)
    delete()
    keyboard.type(interior_comments)
    tab(1)
    enter()
    tab(1)
    keyboard.type('Unknown; Interior not viewed by the appraiser at the request of the lender due to safety and health concerns impacted by the Co-Vid 19 Virus.')
    tab(1)
    enter()
    tab(1)
    keyboard.type('Unknown; Interior not viewed by the appraiser at the request of the lender due to safety and health concerns impacted by the Co-Vid 19 Virus.')
    tab(1)
    # at interior table, first level entrance  
    # auto fill interior table
    tab(1)
    if interior1[3] != '1':
        delete()
        tab(1)
        delete()
        tab(1)
        delete()
        tab(1)
    else:
        tab(3)
    enter()
    #in floor 1bathroom
    tab(1)
    delete()
    table_enter(interior1[25])
    tab(2)
    t_full = int(interior1[21]) + int(interior1[23])
    table_enter(t_full)
    enter()
    tab(1)
    t_bed = int(interior1[17]) + int(interior1[19])
    table_enter(t_bed)
    tab(1)
    delete()
    table_enter(interior1[9])
    tab(1)
    delete()
    if interior1[15] == '1':
        keyboard.type('x')
    tab(13)
    # at 2nd F living
    if storey == '1' or type == '3':
        tab(3)
        enter()
        tab(2)
        delete()
        tab(1)
        delete()
        enter()
        tab(1)
        delete()
        tab(38)
    else:
        if interior2[1] == '1':
            keyboard.type('1')
            tab(1)
            keyboard.type('1')
            tab(1)
            keyboard.type('1')
            tab(1)
        else:
            tab(3)
        enter()
        tab(1)
        table_enter(interior2[23])
        tab(1)
        delete()
        tab(1)
        delete()
        t_full = int(interior2[19]) + int(interior2[21])
        table_enter(t_full)
        enter()
        tab(1)
        delete()
        t_bed = int(interior2[15]) + int(interior2[17])
        table_enter(t_bed)
        tab(2)
        if interior2[13] == '1':
            keyboard.type('x')
        tab(8)
        # at third floor living
        if storey == '2 1/2' or storey == '3':
            tab(3)
            enter()
            tab(1)
            table_enter(interior3[17])
            tab(2)
            t_full = int(interior3[13]) + int(interior3[15])
            table_enter(t_full)
            enter()
            tab(1)
            t_bed = int(interior3[9]) + int(interior3[11])
            table_enter(t_bed)
            tab(2)
            if interior3[7] == '1':
                keyboard.type('x')
            tab(22)
        else:
            tab(28)
    #at basement entrance
    if basement != '3' and type != '3' and basement != '2':
        if all_basement_int[1] == '1':
            keyboard.type('x')
        tab(1)
        table_enter(all_basement_int[3])
        tab(1)
        table_enter(all_basement_int[5])
        tab(1)
        table_enter(all_basement_int[7])
        tab(1)
        enter()
        tab(1)
        table_enter(all_basement_int[23])
        tab(2)
        table_enter(all_basement_int[21])
        enter()
        tab(1)
        table_enter(all_basement_int[19])
        tab(1)
        table_enter(all_basement_int[11])
        tab(1)
        if all_basement_int[15] == '1':
            keyboard.type('x')
        tab(3)
        table_enter(all_basement_int[9])
        tab(13)
    else:
        tab(23)
    time.sleep(12)
    #-----------
    #if dont want auto fill table, use: tab(85)
    # at cost approach table, garage
    if ownership == '0':
        if parking_no == '0':
            keyboard.type('None')
        else:
            keyboard.type(parking_no + ' car garage')
        tab(4)
        if basement == '0':
            keyboard.type('Fully finished')
        elif basement == '1':
            keyboard.type('Partly finished')
        elif basement == '2':
            keyboard.type('Unfinished')
        elif basement == '3':
            keyboard.type('None')
        tab(6)
        delete()
        tab(7)
        delete()
        keyboard.type(depreciation_table(e_age))
    else:
        tab(17)
    tab(7)
    # at page 5- date of sale of subject
    time.sleep(15)
    tab(4)
    keyboard.type(year_built+ '/' + e_age)
    tab(1)
    if e_age == '0E' or e_age == '1E' or e_age == '2E' or e_age == '3E' or e_age =='5E':
        keyboard.type('Good')
    else:
        keyboard.type('Average')
    tab(3)
    # at LOCATION 1 box
    if ownership == '0':
        keyboard.type('LOCATION')
        tab(1)
        if type == '2':
            if townhouse_location == 'interior':
                keyboard.type('Interior unit')
            elif townhouse_location == 'end':
                keyboard.type('End unit')
        else: 
            if street_type == 'residential street':
                keyboard.type('Interior street')
            elif street_type == 'main road':
                keyboard.type('Main road')
            elif street_type == 'commuter street':
                keyboard.type('Commuter street')
        tab(1)
        keyboard.type('LOCATION')
        tab(1)
        if freehold_location[2] == 'other residential lots':
            keyboard.type('Back residential')
        tab(1)
        keyboard.type('INTERIOR FINISHES')
    elif ownership == '1' and (type =='2' or type == '4'):
        delete()
        keyboard.type('LOCATION')
        tab(1)
        if condo_location[3] == "interior":
            keyboard.type('Interior unit')
        elif condo_location[3] == "end":
            keyboard.type('End unit')
        tab(1)
        keyboard.type('LOCATION')
        tab(1)
        keyboard.type('Back residential')
        tab(1)
        keyboard.type('INTERIOR FINISHES')
    else:
        tab(4)

    tab(46)
    # GLA adjustment
    enter()
    keyboard.type(str(round(225*(1-int(depreciation_table(e_age))/100))))
    enter()
    tab(29)
    delete()
    keyboard.type('Please see addendum.')
    tab(1)
    if lease == 'n':
        enter()
        tab(2)
    else:
        tab(1)
        delete()
        keyboard.type('Please see addendum.')
        tab(1)
    right(1)
    tab(1)
    right(1)
    tab(1)
    delete()
    #oversize sale history comment
    if sales_history_length > 4:
        keyboard.type('Please see addendum.')
    else:
        keyboard.type(sales_history_comments)
    tab(6)
    keyboard.type(str(datetime.date(datetime.now())))
    tab(6)
    keyboard.type(str(datetime.date(datetime.now())))
    time.sleep(18)
    # start from page 5, final date
    tab(79)
    time.sleep(1)
    tab(13)
    time.sleep(1)
    tab(1)
    time.sleep(1)
    enter()
    tab(3)
    right(1)
    tab(2)
    right(1)
    tab(6)
    right(2)
    tab(1)
    delete()
    if city.title() == 'Ottawa':
        keyboard.type('1601-22')
    else:
        keyboard.type('1244-22')
    tab(1)
    keyboard.type(str(datetime.date(datetime.now())))
    tab(1)
    delete()
    keyboard.type('not inspected')
    # at page 8 - view property (date) -NOT INSPECTED
    time.sleep(1)
    tab(3)
    delete()
    tab(4)
    delete()
    tab(1)
    time.sleep(1)
    # at page9 
    tab(7)
    time.sleep(1)
    # at page 10 first image
    tab(1)
    down(11)
    tab(2)
    down(14)
    tab(4)
    down(12)
    tab(5)
    time.sleep(1)
    tab(1)
    time.sleep(1)
    tab(1)
    enter()
    tab(1)
    enter()
    tab(2)
    enter()
    tab(1)
    enter()
    tab(2)
    enter()
    tab(8)
    time.sleep(1)
    tab(8)
    time.sleep(1)
    tab(1)
    time.sleep(1)
    tab(4)
    if ownership_restriction_checkbox == 'y':
        tab(1)
    right(1)
    # finishes at page 3- highest use checkbox

#---TRASH CODE-----
# testing variables, not useful function
def assign_var():
    global city, district, ownership, type, condo_location, occupy, adverse, facilities, townhouse_location, adverse_range, nuclear_station, street_type
    global storey, electric, parking_type, parking_no, lease, basement, freehold_location, total_bed, total_partial, total_bath, interior_finishes, basement_kitchen
    global basement_comments, interior_comments, sales_history_comments, site_neighbourhood_comments, year_built, sqft, exterior_finish, ownership_restriction_checkbox, interior1_floor, interior2_floor, interior3_floor
    city = 'Ottawa'
    district = 'Markville'
    # Free=0/condo=1
    ownership = '1'
    # Det=0/Semi=1/Town=2/Apt=3/Stack=4
    type = '3'
    storey = '2'
    # lot shape, interior/end/corner, backing onto:
    freehold_location = 'rectangular,interior,other residential lots'
    # townhouse location
    townhouse_location = 'end'
    # Commercial=0/Residential=1/Main=2/Commuter=3 front street
    street_type = '1'
    # occupy Owner=0/Tenant=1/Both=2/vacant=3
    occupy = '0'
    # Attach=0/Builtin=1/Detach=2/Underg=3/Aboveg=4/None+Driveway=5/NoneNone=6
    parking_type = '3'
    parking_no = '1'
    # finished=0/partly=1/unfin=2/none=3
    basement = '0'
    basement_kitchen = 0
    adverse = ''
    electric = '1'
    total_bed = '3'
    total_partial = '1' 
    total_bath = '2'
    # potlight/crown moulding
    interior_finishes = '1,1'
    # brick=0/stone=1/vinyl=2/concrete=3/stucco=4 (a list)
    exterior_finish = ['0', '2']
    basement_comments = 'dummy for basement comment'
    interior_comments = 'dummy for interior comment'
    sales_history_comments = 'dummy for sales history'
    site_neighbourhood_comments = 'dummy for site&neighbourhood'
    year_built = '2001'
    sqft = '1500'
    interior1_floor, interior2_floor, interior3_floor = ['2','3'], ['1'], ['3']
    ownership_restriction_checkbox = 'n'

# OUTDATED Replaced by input_excel()
# get property information
def takes_info():
    global city, district, ownership, type, condo_location, occupy, adverse, facilities, townhouse_location, adverse_range, nuclear_station, street_type, storey, electric, parking_type, parking_no, lease, basement, freehold_location, ownership_restriction_checkbox
    print ('City?')
    city = input().title()
    print ('District')
    district = input().title()
    # type of property
    print ('Free=0/condo=1?')
    ownership = input()
    while True:
        print ('property type Det=0/Semi=1/Town=2/Apt=3/Stack=4?')
        type = input()
        if type == '0' and ownership == '1' or type == '1' and ownership =='1' or type =='3' and ownership =='0' or type == '4' and ownership == '0':
            print ('Invalid input')
        else:
            break
    #no. of storey
    if type != '3':
        print ('How many storey?')
        storey = input()
    else:
        storey = '1'
    if ownership == '1':
        while True:
            print ('Quadrant, Street1, Street2, end/corner/interior')
            condo_location = input().split(',')
            if len(condo_location) == 4:
                break
            else:
                print ('invalid input')
    if ownership == '0':
        while True:
            print ('lot shape, interior/end/corner, backing onto:')
            freehold_location = input().split(',')
            if len(freehold_location) == 3:
                break
            else:
                print ('invalid input')
    if ownership =='0' and type == '2':
        print ('townhouse location interior=0/ end = 1?')
        townhouse_location = input()
        if townhouse_location == '0':
            townhouse_location = 'interior'
        elif townhouse_location == '1':
            townhouse_location = 'end'
    #street type
    print ('Commercial=0/Residential=1/Main=2/Commuter=3 front street')
    street_type = input()
    if street_type == '0':
        street_type = 'commercial street'
    elif street_type == '1':
        street_type = 'residential street'
    elif street_type == '2':
        street_type = 'main road'
    elif street_type == '3':
        street_type = 'commuter street'
    # occupied?
    print ('Owner=0/Tenant=1/Both=2/vacant=3')
    occupy = input()
    # parking
    print ('Attach=0/Builtin=1/Detach=2/Underg=3/Aboveg=4/None+Driveway=5/NoneNone=6')
    parking_type = input ()
    print ('How many parking space?')
    parking_no = input()
    # basement
    if type != '3':
        print('finished=0/partly=1/unfin=2/none=3')
        basement = input()
    # adverse or no
    print ('airport=0/railway=1/main=2/commuter=3/hydrolane=4/nuclear=5' )
    adverse = input().split(',')
    # ask for range if have airport/railway/lane/nuclear
    adverse_range = []
    if len(adverse[0]) != 0:
        for i in adverse:
            if i =='0':
                print ('Airport range(KM)?')
                adverse_range.append(input())
            elif i =='1':
                print ('Railway range(M)?')
                adverse_range.append(input())
            elif i == '2' or i == '3':
                adverse_range.append('')
            elif i =='4':
                print ('Hydro lane range(M)?')
                adverse_range.append(input())
            elif i =='5':
                print ('Nuclear range(KM)?')
                adverse_range.append(input())
                print ('Pickering=0/Darlington=1')
                nuclear_station = input()
    print ('Nearby park/school')
    facilities = input()
    print ('electric underground=0/overhead=1')
    electric = input()
    if electric == '0':
        electric = 'underground'
    else:
        electric = 'overhead'
    #lease required?
    print ('leased required? y/n')
    lease = input()
    print ('extra ownership check box exist? y/n')
    ownership_restriction_checkbox = input()

#---TRASH ^-----

def full_report():
    result = ''
    print('Welcome')
    #assign_var()
    #import excel sheet
    input_excel()
    # generate long comments
    basement_comments = (basement_comment_gen())
    interior_comments = (interior_info())
    sales_history_comments = (sales_history())
    site_neighbourhood_comments = (neighbour_site())
    result += basement_comments + interior_comments + ("Sales History\n") +sales_history_comments + site_neighbourhood_comments
    text_file = open("Output.txt", "w")
    text_file.write(result)
    text_file.close()
    # auto input data into pdf
    auto_input(basement_comments, interior_comments, sales_history_comments)

def sales_history_only():
    input_excel()
    print ("Sales History\n" + sales_history())

def main():
    global keyboard
    keyboard = Controller()
    full_report()
    #sales_history_only()
    print ('done')



if __name__ == "__main__":
    main()

